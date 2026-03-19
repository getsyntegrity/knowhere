import io
import os
import re
import uuid
import datetime
import threading
import numpy as np
import pandas as pd
import openpyxl

from collections import OrderedDict
from typing import List, Union, Dict, Tuple, Optional
from shared.core.config import settings
from shared.utils.OpenAICompatibleClientSync import get_openai_client
from shared.services.ai.prompt_service import build_prompt
from shared.services.ai.response_process_service import eval_response
from app.services.common.kb_utils import (flatten_dic2paths, gen_str_codes,
                                          get_str_time, process_dup_paths_df,
                                          remove_spaces)
from shared.utils.text_utils import tokenize2stw_remove, remove_duplicates_orderkept

from app.services.document_parser.txt_parser import extract_summary_keywords
from app.services.document_parser.html_parser import df2html, tb_htmlstr_to_df, html_to_md_lines
from shared.utils.CommonHelperSync import load_file_bytes
from bs4 import BeautifulSoup
from loguru import logger

from shared.core.exceptions.domain_exceptions import TableParsingException
from shared.core.exceptions.knowhere_exception import KnowhereException

g_tbl_lock = threading.Lock()

# ============================================================================
# PRECISION MODE: Excel Header Detection with Merge Cell Metadata
# ============================================================================

def _get_merged_cell_value(ws, row: int, col: int, merged_ranges: list):
    """
    Get the value of a cell, accounting for merged cell regions.
    For merged cells, returns the value from the top-left corner of the merge range.
    
    Args:
        ws: openpyxl worksheet
        row: 1-indexed row number
        col: 1-indexed column number
        merged_ranges: list of merged cell ranges from ws.merged_cells.ranges
    
    Returns:
        The cell value (from merge origin if applicable)
    """
    for mr in merged_ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            # This cell is part of a merged region, get value from top-left
            return ws.cell(mr.min_row, mr.min_col).value
    # Not a merged cell, return direct value
    return ws.cell(row, col).value


# ============================================================================
# NEW: Enhanced Header Detection with Row/Column MultiIndex Support
# ============================================================================

# Data types that indicate a cell is data, not header (parameterized for future extension)
DATA_TYPES_TO_EXCLUDE = (int, float, datetime.datetime)


def _get_unique_cells_in_row(ws, row: int, col_range: Tuple[int, int], merged_ranges: list) -> List[dict]:
    """Get all unique cells in a row, treating merged cells as single cells.
    
    Returns: List of {col_start, col_end, value, is_merged}
    """
    c_start, c_end = col_range
    cells = []
    visited_cols = set()
    
    for col in range(c_start, c_end + 1):
        if col in visited_cols:
            continue
        
        in_merge = False
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                val = ws.cell(mr.min_row, mr.min_col).value
                merge_col_end = min(mr.max_col, c_end)
                
                for mc in range(mr.min_col, merge_col_end + 1):
                    visited_cols.add(mc)
                
                cells.append({
                    'col_start': mr.min_col,
                    'col_end': merge_col_end,
                    'value': val,
                    'is_merged': True
                })
                in_merge = True
                break
        
        if not in_merge:
            val = ws.cell(row, col).value
            cells.append({
                'col_start': col,
                'col_end': col,
                'value': val,
                'is_merged': False
            })
            visited_cols.add(col)
    
    return cells


def _get_unique_cells_in_col(ws, col: int, row_range: Tuple[int, int], merged_ranges: list) -> List[dict]:
    """Get all unique cells in a column, treating merged cells as single cells."""
    r_start, r_end = row_range
    cells = []
    visited_rows = set()
    
    for row in range(r_start, r_end + 1):
        if row in visited_rows:
            continue
        
        in_merge = False
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                val = ws.cell(mr.min_row, mr.min_col).value
                merge_row_end = min(mr.max_row, r_end)
                
                for mr_row in range(mr.min_row, merge_row_end + 1):
                    visited_rows.add(mr_row)
                
                cells.append({
                    'row_start': mr.min_row,
                    'row_end': merge_row_end,
                    'value': val,
                    'is_merged': True
                })
                in_merge = True
                break
        
        if not in_merge:
            val = ws.cell(row, col).value
            cells.append({
                'row_start': row,
                'row_end': row,
                'value': val,
                'is_merged': False
            })
            visited_rows.add(row)
    
    return cells


def _is_candidate_header_row(
    ws, row: int, col_range: Tuple[int, int], merged_ranges: list,
    exclude_types: tuple = DATA_TYPES_TO_EXCLUDE
) -> bool:
    """Check if a row is a candidate header row.
    
    Logic: Row is a candidate if all cells are text (no numbers/dates).
    Merged cells are treated as single cells.
    """
    cells = _get_unique_cells_in_row(ws, row, col_range, merged_ranges)
    
    has_any_value = False
    for cell in cells:
        val = cell['value']
        if val is None:
            continue
        has_any_value = True
        
        if isinstance(val, bool):
            continue
        if isinstance(val, exclude_types):
            return False
    
    return has_any_value


def _is_candidate_header_col(
    ws, col: int, row_range: Tuple[int, int], merged_ranges: list,
    exclude_types: tuple = DATA_TYPES_TO_EXCLUDE
) -> bool:
    """Check if a column is a candidate header column (for row index)."""
    cells = _get_unique_cells_in_col(ws, col, row_range, merged_ranges)
    
    has_any_value = False
    for cell in cells:
        val = cell['value']
        if val is None:
            continue
        has_any_value = True
        
        if isinstance(val, bool):
            continue
        if isinstance(val, exclude_types):
            return False
    
    return has_any_value


def _detect_header_regions(
    ws, row_range: Tuple[int, int], col_range: Tuple[int, int], merged_ranges: list
) -> Tuple[List[int], List[int]]:
    """Detect header rows and columns.
    
    Scans rows first, then scans columns only in the data region (excluding header rows).
    This prevents header row content from influencing column header detection.
    
    Returns:
        header_rows: List of candidate header row numbers (1-indexed)
        header_cols: List of candidate header column numbers (1-indexed)
    """
    r_start, r_end = row_range
    c_start, c_end = col_range
    
    # Scan for candidate header rows (top to bottom)
    header_rows = []
    for row in range(r_start, r_end + 1):
        if _is_candidate_header_row(ws, row, col_range, merged_ranges):
            header_rows.append(row)
        else:
            break
    
    # Determine data region (excluding header rows)
    data_row_start = header_rows[-1] + 1 if header_rows else r_start
    
    # Skip column scanning if no data rows remain
    if data_row_start > r_end:
        return header_rows, []
    
    # Scan for candidate header columns (left to right) - only in data region
    header_cols = []
    data_row_range = (data_row_start, r_end)
    for col in range(c_start, c_end + 1):
        if _is_candidate_header_col(ws, col, data_row_range, merged_ranges):
            header_cols.append(col)
        else:
            break
    
    return header_rows, header_cols


def _build_column_multiindex(
    ws, header_rows: List[int], col_range: Tuple[int, int], merged_ranges: list
) -> Union[pd.Index, pd.MultiIndex]:
    """Build column MultiIndex from header rows."""
    c_start, c_end = col_range
    levels = []
    
    for row in header_rows:
        row_values = []
        for col in range(c_start, c_end + 1):
            val = _get_merged_cell_value(ws, row, col, merged_ranges)
            row_values.append(str(val).strip() if val else '')
        levels.append(row_values)
    
    # Forward fill for merged cells
    for idx, level in enumerate(levels):
        filled = []
        last = ''
        for val in level:
            if val:
                last = val
            filled.append(last if last else val)
        levels[idx] = filled
    
    if len(levels) == 1:
        return pd.Index(levels[0])
    return pd.MultiIndex.from_arrays(levels)


def _build_row_multiindex(
    ws, header_cols: List[int], row_range: Tuple[int, int], merged_ranges: list,
    header_rows: List[int] = None
) -> Union[pd.Index, pd.MultiIndex]:
    """Build row MultiIndex from header columns.
    
    Args:
        header_rows: If provided, use the last header row's values as index names
    """
    r_start, r_end = row_range
    levels = []
    names = []
    
    for col in header_cols:
        col_values = []
        for row in range(r_start, r_end + 1):
            val = _get_merged_cell_value(ws, row, col, merged_ranges)
            col_values.append(str(val).strip() if val else '')
        levels.append(col_values)
        
        # Get the column name from the last header row
        if header_rows:
            name_row = header_rows[-1]
            name_val = _get_merged_cell_value(ws, name_row, col, merged_ranges)
            names.append(str(name_val).strip() if name_val else None)
        else:
            names.append(None)
    
    # Forward fill for merged cells
    for idx, level in enumerate(levels):
        filled = []
        last = ''
        for val in level:
            if val:
                last = val
            filled.append(last if last else val)
        levels[idx] = filled
    
    if len(levels) == 1:
        idx = pd.Index(levels[0])
        idx.name = names[0] if names else None
        return idx
    return pd.MultiIndex.from_arrays(levels, names=names)


def _parse_subtable(
    ws, row_range: Tuple[int, int], col_range: Tuple[int, int], merged_ranges: list
) -> dict:
    """Parse a subtable with new header detection logic.
    
    Returns:
        dict with keys: df, header_rows, header_cols, fallback_col_header, fallback_row_header
    """
    r_start, r_end = row_range
    c_start, c_end = col_range
    
    header_rows, header_cols = _detect_header_regions(ws, row_range, col_range, merged_ranges)
    
    total_rows = r_end - r_start + 1
    total_cols = c_end - c_start + 1
    
    # Fall-back check: if all rows/cols are headers, treat as no-header
    fallback_col_header = len(header_rows) == total_rows
    fallback_row_header = len(header_cols) == total_cols
    
    # Determine data region
    if fallback_col_header:
        data_row_start = r_start
        columns = None
    else:
        data_row_start = header_rows[-1] + 1 if header_rows else r_start
        columns = _build_column_multiindex(ws, header_rows, col_range, merged_ranges) if header_rows else None
    
    if fallback_row_header:
        data_col_start = c_start
        row_index = None
    else:
        data_col_start = header_cols[-1] + 1 if header_cols else c_start
        row_index = _build_row_multiindex(ws, header_cols, (data_row_start, r_end), merged_ranges, header_rows) if header_cols else None
    
    # Read data
    data = []
    for row in range(data_row_start, r_end + 1):
        row_data = []
        for col in range(data_col_start, c_end + 1):
            val = _get_merged_cell_value(ws, row, col, merged_ranges)
            row_data.append(val)
        data.append(row_data)
    
    # Adjust column index if there are row index columns
    if columns is not None and header_cols and not fallback_row_header:
        if isinstance(columns, pd.MultiIndex):
            columns = columns[len(header_cols):]
        else:
            columns = columns[len(header_cols):]
    
    df = pd.DataFrame(data, columns=columns, index=row_index)
    
    # Append original Excel row numbers as the last column for cross-referencing
    excel_row_numbers = list(range(data_row_start, r_end + 1))
    if isinstance(df.columns, pd.MultiIndex):
        n_levels = df.columns.nlevels
        src_row_key = tuple(['_src_row'] + [''] * (n_levels - 1))
        df[src_row_key] = excel_row_numbers
    else:
        df['_src_row'] = excel_row_numbers
    
    return {
        'df': df,
        'header_rows': header_rows if not fallback_col_header else [],
        'header_cols': header_cols if not fallback_row_header else [],
        'fallback_col_header': fallback_col_header,
        'fallback_row_header': fallback_row_header,
    }


# ============================================================================
# Sheet Splitting: Detect true separators and split into subtables
# ============================================================================

def _find_effective_range(ws, row_range: Tuple[int, int], col_range: Tuple[int, int]) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    """Find the effective (non-empty) row and column ranges within a region."""
    r_start, r_end = row_range
    c_start, c_end = col_range
    
    eff_r_start, eff_r_end = None, None
    eff_c_start, eff_c_end = None, None
    
    for row in range(r_start, r_end + 1):
        for col in range(c_start, c_end + 1):
            if ws.cell(row, col).value is not None:
                if eff_r_start is None:
                    eff_r_start = row
                eff_r_end = row
                if eff_c_start is None or col < eff_c_start:
                    eff_c_start = col
                if eff_c_end is None or col > eff_c_end:
                    eff_c_end = col
    
    if eff_r_start is None:
        return ((r_start, r_start), (c_start, c_start))
    
    return ((eff_r_start, eff_r_end), (eff_c_start, eff_c_end))


def _is_true_separator_row(ws, row: int, effective_col_range: Tuple[int, int], merged_ranges: list = None) -> bool:
    """Check if a row is a true separator (all empty within effective column range).
    
    Considers merged cells - a cell is not empty if it's part of any merged range.
    """
    c_start, c_end = effective_col_range
    merged_ranges = merged_ranges or []
    
    for col in range(c_start, c_end + 1):
        # Check if cell has a value
        if ws.cell(row, col).value is not None:
            return False
        # Check if cell is part of a merged range
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return False  # Part of a merge, not truly empty
    return True


def _is_true_separator_col(ws, col: int, effective_row_range: Tuple[int, int], merged_ranges: list = None) -> bool:
    """Check if a column is a true separator (all empty within effective row range).
    
    Considers merged cells - a cell is not empty if it's part of any merged range.
    """
    r_start, r_end = effective_row_range
    merged_ranges = merged_ranges or []
    
    for row in range(r_start, r_end + 1):
        # Check if cell has a value
        if ws.cell(row, col).value is not None:
            return False
        # Check if cell is part of a merged range
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return False  # Part of a merge, not truly empty
    return True


def _find_separator_groups(items: List[int]) -> List[List[int]]:
    """Group consecutive separator items together."""
    if not items:
        return []
    
    groups = []
    current_group = [items[0]]
    
    for i in range(1, len(items)):
        if items[i] == items[i-1] + 1:
            current_group.append(items[i])
        else:
            groups.append(current_group)
            current_group = [items[i]]
    
    groups.append(current_group)
    return groups


def _split_sheet_recursive(
    ws, 
    row_range: Tuple[int, int], 
    col_range: Tuple[int, int],
    merged_ranges: list = None,
    min_rows: int = 2,
    min_cols: int = 2
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    """
    Recursively split a sheet region into subtables based on true separators.
    
    Args:
        merged_ranges: List of merged cell ranges to consider when detecting separators
    
    Returns list of (row_range, col_range) tuples for each subtable.
    """
    r_start, r_end = row_range
    c_start, c_end = col_range
    merged_ranges = merged_ranges or []
    
    # Find effective range (trim empty edges)
    (eff_r_start, eff_r_end), (eff_c_start, eff_c_end) = _find_effective_range(ws, row_range, col_range)
    
    # If region is too small or empty, return as-is or empty
    if eff_r_end - eff_r_start + 1 < min_rows or eff_c_end - eff_c_start + 1 < min_cols:
        if eff_r_start is not None:
            return [((eff_r_start, eff_r_end), (eff_c_start, eff_c_end))]
        return []
    
    # Find true separator rows (considering merged cells)
    separator_rows = []
    for row in range(eff_r_start + 1, eff_r_end):
        if _is_true_separator_row(ws, row, (eff_c_start, eff_c_end), merged_ranges):
            separator_rows.append(row)
    
    # Find true separator columns (considering merged cells)
    separator_cols = []
    for col in range(eff_c_start + 1, eff_c_end):
        if _is_true_separator_col(ws, col, (eff_r_start, eff_r_end), merged_ranges):
            separator_cols.append(col)
    
    # Group consecutive separators
    row_groups = _find_separator_groups(separator_rows)
    col_groups = _find_separator_groups(separator_cols)
    
    # Choose split direction
    do_row_split = len(row_groups) > 0 and (len(col_groups) == 0 or len(row_groups) <= len(col_groups))
    do_col_split = len(col_groups) > 0 and not do_row_split
    
    if do_row_split:
        subtables = []
        prev_end = eff_r_start
        for group in row_groups:
            if group[0] > prev_end:
                sub_result = _split_sheet_recursive(ws, (prev_end, group[0] - 1), (eff_c_start, eff_c_end), merged_ranges, min_rows, min_cols)
                subtables.extend(sub_result)
            prev_end = group[-1] + 1
        if prev_end <= eff_r_end:
            sub_result = _split_sheet_recursive(ws, (prev_end, eff_r_end), (eff_c_start, eff_c_end), merged_ranges, min_rows, min_cols)
            subtables.extend(sub_result)
        return subtables
    
    elif do_col_split:
        subtables = []
        prev_end = eff_c_start
        for group in col_groups:
            if group[0] > prev_end:
                sub_result = _split_sheet_recursive(ws, (eff_r_start, eff_r_end), (prev_end, group[0] - 1), merged_ranges, min_rows, min_cols)
                subtables.extend(sub_result)
            prev_end = group[-1] + 1
        if prev_end <= eff_c_end:
            sub_result = _split_sheet_recursive(ws, (eff_r_start, eff_r_end), (prev_end, eff_c_end), merged_ranges, min_rows, min_cols)
            subtables.extend(sub_result)
        return subtables
    
    else:
        return [((eff_r_start, eff_r_end), (eff_c_start, eff_c_end))]


# ============================================================================
# Post-split Merge: Absorb small fragments into nearest neighbor
# ============================================================================

def _count_non_empty_cells(ws, row_range: Tuple[int, int], col_range: Tuple[int, int]) -> int:
    """Count non-empty cells in a region."""
    count = 0
    for r in range(row_range[0], row_range[1] + 1):
        for c in range(col_range[0], col_range[1] + 1):
            if ws.cell(r, c).value is not None:
                count += 1
    return count


def _merge_small_subtables(
    ws,
    subtables: List[Tuple[Tuple[int, int], Tuple[int, int]]],
    min_cells: int = 4
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    """Merge subtables that have too few non-empty cells into their nearest neighbor.

    This is a post-processing step after _split_sheet_recursive to prevent
    over-fragmentation. Fragments with fewer than min_cells non-empty cells
    are iteratively absorbed into the nearest neighbor subtable (by bounding-box
    distance), expanding the neighbor's bounding box to encompass both regions.

    Args:
        ws: openpyxl worksheet
        subtables: list of (row_range, col_range) tuples from _split_sheet_recursive
        min_cells: minimum non-empty cells for a subtable to be kept standalone

    Returns:
        Merged list of (row_range, col_range) tuples
    """
    if len(subtables) <= 1:
        return subtables

    # Build working list with cell counts
    items = []
    for (rr, cr) in subtables:
        count = _count_non_empty_cells(ws, rr, cr)
        items.append({'rr': rr, 'cr': cr, 'cells': count})

    # Iteratively merge the smallest sub-threshold fragment
    changed = True
    while changed and len(items) > 1:
        changed = False

        # Find the smallest fragment below threshold
        min_idx = None
        for i, item in enumerate(items):
            if item['cells'] < min_cells:
                if min_idx is None or item['cells'] < items[min_idx]['cells']:
                    min_idx = i

        if min_idx is None:
            break  # All subtables are above threshold

        # Find nearest neighbor by bounding-box gap distance
        src = items[min_idx]
        best_j = None
        best_dist = float('inf')
        for j, tgt in enumerate(items):
            if j == min_idx:
                continue
            row_gap = max(0, tgt['rr'][0] - src['rr'][1] - 1, src['rr'][0] - tgt['rr'][1] - 1)
            col_gap = max(0, tgt['cr'][0] - src['cr'][1] - 1, src['cr'][0] - tgt['cr'][1] - 1)
            dist = row_gap + col_gap
            if dist < best_dist or (dist == best_dist and tgt['cells'] > items[best_j]['cells']):
                best_dist = dist
                best_j = j

        if best_j is None:
            break  # Should not happen when len(items) > 1

        # Merge: expand neighbor's bounding box to encompass both
        tgt = items[best_j]
        merged_rr = (min(src['rr'][0], tgt['rr'][0]), max(src['rr'][1], tgt['rr'][1]))
        merged_cr = (min(src['cr'][0], tgt['cr'][0]), max(src['cr'][1], tgt['cr'][1]))
        items[best_j] = {
            'rr': merged_rr,
            'cr': merged_cr,
            'cells': src['cells'] + tgt['cells'],
        }

        logger.debug(
            f"Merged small fragment (rows={src['rr']}, cols={src['cr']}, "
            f"cells={src['cells']}) into neighbor (rows={tgt['rr']}, cols={tgt['cr']})"
        )

        del items[min_idx]
        changed = True

    return [(item['rr'], item['cr']) for item in items]


def parse_headers_from_excel(
    file_source: Union[str, io.BytesIO],
    sheet_name: Optional[str] = None,
    split_subtables: bool = True,
    include_hidden_sheets: bool = False
) -> Dict[str, pd.DataFrame]:
    """
    Parse Excel file using openpyxl to accurately detect headers via merged cell metadata.
    
    This is the PRECISION MODE for Excel parsing - it uses the actual merge cell
    information from the Excel file to build correct MultiIndex headers without
    relying on LLM or heuristics.
    
    Args:
        file_source: Path to Excel file or BytesIO stream
        sheet_name: Specific sheet to parse (None = all sheets)
        split_subtables: If True, split sheets into subtables based on empty row/column separators (default: True)
        include_hidden_sheets: If True, parse hidden/very-hidden sheets. Default False (skip them).
    
    Returns:
        Dictionary mapping sheet/subtable names to DataFrames with correctly set headers
        When split_subtables=True, keys are like 'SheetName', 'SheetName_2', 'SheetName_3' etc.
    """
    try:
        # Load workbook with data_only=True to get calculated values
        if isinstance(file_source, str):
            wb = openpyxl.load_workbook(file_source, data_only=True)
        else:
            # BytesIO stream
            file_source.seek(0)  # Ensure we're at the start
            wb = openpyxl.load_workbook(file_source, data_only=True)
        
        results = {}
        sheets_to_parse = [sheet_name] if sheet_name else wb.sheetnames
        
        for sn in sheets_to_parse:
            if sn not in wb.sheetnames:
                logger.warning(f"Sheet '{sn}' not found in workbook, skipping")
                continue
            
            ws = wb[sn]
            
            # Skip hidden sheets unless explicitly included
            if not include_hidden_sheets and ws.sheet_state != 'visible':
                logger.info(f"Sheet '{sn}' is hidden (state={ws.sheet_state}), skipping")
                continue
            
            # Skip empty sheets
            if ws.max_row is None or ws.max_row == 0:
                logger.debug(f"Sheet '{sn}' is empty, skipping")
                continue
            
            # Get merged cell ranges
            merged_ranges = list(ws.merged_cells.ranges)
            logger.debug(f"Sheet '{sn}': found {len(merged_ranges)} merged cell ranges")
            
            if split_subtables:
                # Split sheet into subtables (considers merged cells)
                subtable_regions = _split_sheet_recursive(
                    ws, (1, ws.max_row), (1, ws.max_column or 1), merged_ranges
                )
                # Merge back small fragments to prevent over-fragmentation
                before_count = len(subtable_regions)
                subtable_regions = _merge_small_subtables(ws, subtable_regions)
                if len(subtable_regions) != before_count:
                    logger.info(
                        f"Sheet '{sn}': merged {before_count} subtables → {len(subtable_regions)} "
                        f"(absorbed {before_count - len(subtable_regions)} small fragments)"
                    )
                logger.debug(f"Sheet '{sn}': {len(subtable_regions)} subtables after merge")
                
                for idx, (row_range, col_range) in enumerate(subtable_regions):
                    result = _parse_subtable(ws, row_range, col_range, merged_ranges)
                    df = result['df']
                    
                    # Store header_cols count in DataFrame attrs for later use in HTML rendering
                    df.attrs['row_header_cols'] = len(result['header_cols'])
                    
                    # Generate unique key for each subtable
                    if idx == 0:
                        key = sn
                    else:
                        key = f"{sn}_{idx + 1}"
                    
                    logger.debug(
                        f"Subtable '{key}': rows={row_range}, cols={col_range}, "
                        f"header_rows={result['header_rows']}, header_cols={result['header_cols']}"
                    )
                    
                    results[key] = df
            else:
                # Treat entire sheet as one subtable
                row_range = (1, ws.max_row)
                col_range = (1, ws.max_column or 1)
                
                result = _parse_subtable(ws, row_range, col_range, merged_ranges)
                df = result['df']
                
                # Store header_cols count in DataFrame attrs for later use in HTML rendering
                df.attrs['row_header_cols'] = len(result['header_cols'])
                
                logger.debug(
                    f"Sheet '{sn}': header_rows={result['header_rows']}, "
                    f"header_cols={result['header_cols']}, "
                    f"fallback_col={result['fallback_col_header']}, "
                    f"fallback_row={result['fallback_row_header']}"
                )
                
                results[sn] = df
        
        wb.close()
        return results
        
    except Exception as e:
        logger.error(f"Error parsing Excel with precision mode: {e}")
        raise TableParsingException(
            user_message="Failed to parse Excel file headers",
            reason="EXCEL_PRECISION_PARSE_FAILED",
            internal_message=str(e),
            original_exception=e
        )


def identify_tables(line):
    """Identify if a line contains a table.
    
    Note: For HTML tables, use merge_html_tables() from html_parser.py 
    to preprocess multi-line tables before calling this function.
    """
    # HTML table: complete <table>...</table> in one line
    html_tb_pattern = r'<table.*?>.*?</table>'
    tables = re.findall(html_tb_pattern, line, re.DOTALL)
    if bool(tables):
        return True, 'html', tables
    
    # MD table: lines starting and ending with |
    if line.startswith('|') and line.endswith('|'):
        return True, 'md', []
    
    return False, None, None


def df2md(tb_df: pd.DataFrame,
    *,
    index: bool = False,
    na_rep: str = "—"
    ) -> str:
    """Convert DataFrame to Markdown table format with dynamic column widths.
    
    Note: Truncation should be done externally using truncate_text before calling this function.
    
    Args:
        tb_df: Input DataFrame
        index: Whether to include index column
        na_rep: String to represent NA values
    
    Returns:
        Markdown table string
    """
    import unicodedata
    def get_display_width(text: str) -> int:
        """eval width for both ASCII and Chinese"""
        width = 0
        for char in text:
            if unicodedata.east_asian_width(char) in ('F', 'W'):
                width += 2
            else:
                width += 1
        return width
    
    def pad_to_width(text: str, target_width: int) -> str:
        current_width = get_display_width(text)
        padding = target_width - current_width
        return text + ' ' * max(0, padding)
    
    df = tb_df.copy()
    
    # Handle index
    if index:
        df = df.reset_index()
    
    # Replace NA values
    df = df.fillna(na_rep)
    
    # Convert all values to string
    df = df.astype(str)
    
    # Calculate column widths based on actual display width (no truncation)
    col_widths = {}
    for col in df.columns:
        header_width = get_display_width(str(col))
        max_content_width = max(df[col].apply(get_display_width)) if len(df) > 0 else 0
        col_widths[col] = max(header_width, max_content_width)
    
    # Build header row
    header_cells = [pad_to_width(str(col), col_widths[col]) for col in df.columns]
    header_line = "| " + " | ".join(header_cells) + " |"
    
    # Build separator row
    separator_cells = ["-" * col_widths[col] for col in df.columns]
    separator_line = "|-" + "-|-".join(separator_cells) + "-|"
    
    # Build data rows
    data_lines = []
    for _, row in df.iterrows():
        cells = [pad_to_width(str(row[col]), col_widths[col]) for col in df.columns]
        data_lines.append("| " + " | ".join(cells) + " |")
    
    # Combine all parts
    lines = [header_line, separator_line] + data_lines
    return "\n".join(lines)


def clean_html_tb(html: str) -> str:
    soup = BeautifulSoup(html, 'html.parser')
    for row in soup.find_all("tr"):
        seen = set()
        unique_cells = []
        for cell in row.find_all("td", recursive=False):
            content = cell.encode_contents()
            if content not in seen:
                seen.add(content)
                unique_cells.append(cell)
        row.clear()
        for cell in unique_cells:
            row.append(cell)
    return soup.prettify()


def extract_tables_by_forms(tb_txt, form):
    if form=='html':
        return tb_txt
    elif form=='md':
        tb_df = pd.read_table(pd.io.common.StringIO(tb_txt), sep='|', engine='python', on_bad_lines='skip')
        tb_df = tb_df.drop(columns=tb_df.columns[0])  # Drop extra leading column
        tb_df = tb_df.drop(columns=tb_df.columns[-1]) # Drop extra trailing column
        tb_df.columns = tb_df.columns.str.strip()  # Clean up headers
        # Filter out MD separator lines (e.g. "---", ":---:", "---:")
        separator_pattern = r'^[\s\-:]+$'
        tb_df = tb_df[~tb_df.apply(lambda row: row.astype(str).str.match(separator_pattern).all(), axis=1)]
        tb_strs = tb_df.to_html(index=False)
    else:
        tb_strs = None # UNDER DEVELOPMENT other forms of tables...
    return tb_strs


def parse_headers(df_temp, paras=None, header_window=5, smart_headers=True):
    def parse_headers_nonsmart(df_):
        non_na_row = df_[df_.notna().any(axis=1)].head(1)
        header_id = non_na_row.index[-1] if not non_na_row.empty else None
        header_rows = list(range(header_id+1))
        return header_rows

    if not pd.isna(df_temp.columns).all(): # If columns are not all NaN, no need to add extra row
        df_temp.loc[-1] = df_temp.columns
        df_temp.index = df_temp.index + 1
        df_temp = df_temp.sort_index()
        df_temp.columns = [np.nan] * df_temp.shape[1]

    if paras['summary_table'] and smart_headers:
        try:
            tb_small = df_temp.head(header_window)
            tb_small_str = df2html(tb_small)
            prompt, temperature, top_p, max_tokens = build_prompt(task="detect-table-headers", texts=tb_small_str, query="", paras=paras)

            messages = [
                {"role": "system", "content": "You are a helpful assistant"},
                {"role": "user", "content": prompt}
            ]

            ctx_task_id = gen_str_codes((str(uuid.uuid4()) + tb_small_str))
            
            # Track task status via Redis (skip in LOCAL_DEBUG mode)
            import os
            if os.getenv("LOCAL_DEBUG", "0") != "1":
                from shared.services.redis.redis_sync_service import SyncRedisServiceFactory
                redis_service = SyncRedisServiceFactory.get_service()
                redis_service.set(f"task:{ctx_task_id}:status", "processing", ttl=7200)
            
            # Use unified AI service
            header_res = get_openai_client().chat_completion(
                messages=messages,
                timeout=60
            )
            header_res = eval_response(header_res)
            # Extract answer field
            if isinstance(header_res, dict):
                answer = header_res.get('answer', [])
            else:
                answer = header_res if isinstance(header_res, list) else []
            
            # Check if answer is empty list
            if not answer or len(answer) == 0:
                logger.warning("AI returned empty list, cannot identify headers, falling back to traditional mode...")
                header_rows = parse_headers_nonsmart(df_temp)
            else:
                try:
                    header_id = answer[-1]
                    header_rows = list(range(header_id + 1))
                except Exception as e:
                    logger.warning(f"Failed to parse header row number: {e}, falling back to traditional mode...")
                    header_rows = parse_headers_nonsmart(df_temp)

        except Exception as e:
            logger.warning(f"Smart header parsing failed: {e}, falling back to traditional mode...")
            header_rows = parse_headers_nonsmart(df_temp)
    else:
        header_rows = parse_headers_nonsmart(df_temp)

    # improve table structure based on header rows
    if len(header_rows)==0 or (all(h is None for h in header_rows)):
        logger.warning("No valid headers detected, fallback to using row 0 as header")
        new_header = df_temp.iloc[0].ffill().bfill().tolist()
        df_temp.columns = new_header
        df_temp = df_temp.iloc[1:].reset_index(drop=True)
        return df_temp
    elif len(header_rows)>1:
        head_lst = []
        for i in range(0, len(header_rows)):
            temp_lst = df_temp.iloc[i].ffill().bfill().tolist()
            head_lst.append(temp_lst)
        new_header = pd.MultiIndex.from_arrays(np.array(head_lst))
    else:
        new_header = df_temp.iloc[header_rows[-1]].ffill().bfill().tolist()

    df_temp.columns = new_header
    df_temp = df_temp.iloc[(header_rows[-1])+1:]
    df_temp = df_temp.reset_index(drop=True)
    return df_temp


def extract_tb_keywords(tb_str, form="html"):
    """Extract keywords from table headers.
    
    TODO: Current implementation has issues with:
    1. colspan/rowspan in HTML - causes column count mismatch when converting to DataFrame
    2. MultiIndex headers - tb_htmlstr_to_df only uses first row as headers, losing hierarchy
    
    For now, consider using _first_cols_rows() from doc_parser.py as a simpler alternative
    that extracts deduplicated first row and first column text.
    """
    if form=='html':
        tb_df = tb_htmlstr_to_df(tb_str)
    else:
        tb_lines = html_to_md_lines(tb_str)
        tb_df = pd.DataFrame(tb_lines)
    tb_keywords = parse_tb_keywords(tb_df)
    return tb_keywords


def parse_tb_keywords(tb_df, kw_spit=">>>"):  # Extract keywords from headers (can also add LLM extraction)
    def parse_single_level_(cols, keywords):
        cols = [str(c) for c in cols]
        for col in cols:
            if kw_spit in col:
                tmp_kw = col.split(">>>")[0]
            else:  # May be first occurrence
                tmp_kw = col
            if tmp_kw not in keywords:
                keywords.append(col)
        keywords_a_level = list(set([k.strip() for k in keywords]))
        return keywords_a_level

    tb_keywords = []
    if isinstance(tb_df.columns, pd.MultiIndex):
        multi_cols = tb_df.columns
        cols_df = pd.DataFrame(multi_cols.tolist(), columns=[f"level_{i}" for i in range(multi_cols.nlevels)])
        for i in range(multi_cols.nlevels):  # Extract each level as list
            level_kws = []
            level_kws = parse_single_level_(cols_df[f"level_{i}"].tolist(), level_kws)
            tb_keywords.extend(level_kws)
    else:
        tb_keywords = parse_single_level_(tb_df.columns, tb_keywords)
    return ';'.join(tb_keywords)


def parse_tb_contents(df_temp, parent_dic=None, file_name='', sheet_name='', row_header_cols=0):
    """Parse table contents and generate HTML.
    
    Args:
        row_header_cols: Number of leftmost columns that are row headers (will be rendered as <th>)
    """
    if parent_dic is None:
        parent_dic = {}

    tb_res = df_temp.fillna('').infer_objects(copy=False)
    tb_strs = df2html(tb_res, row_header_cols=row_header_cols)

    tb_tree = tb_columns_to_tree(df_temp, parent_dic, file_name, sheet_name)
    tb_paths = flatten_dic2paths(tb_tree)
    return tb_paths, tb_strs


def tb_columns_to_tree(df, parent_dic, file_name, sheet_name):
    if isinstance(df.columns, pd.MultiIndex):
        # Convert MultiIndex columns to a nested dictionary (tree-like structure)
        columns = pd.DataFrame(df.columns.tolist())
        for level in range(columns.shape[1]):
            columns[level] = process_duplicate_cols(columns[level])

        new_columns = pd.MultiIndex.from_frame(columns)
        tree_structure = multiindex_to_tree(new_columns)
    else:
        # If columns are not MultiIndex, convert them to a dictionary with empty dictionaries as values
        new_columns = process_duplicate_cols(df.columns)
        tree_structure = {col: {} for col in new_columns}
    
    df.columns = new_columns
    if (not file_name=='') and (not sheet_name==''):
        parent_dic[file_name][sheet_name] = tree_structure
    elif not sheet_name == '':
        parent_dic[sheet_name] = tree_structure
    elif not file_name == '':
        parent_dic[file_name] = tree_structure
    else:
        parent_dic = tree_structure
    return parent_dic
        

def multiindex_to_tree(multiindex):
    """ Convert a MultiIndex to a tree-like nested dictionary structure. """
    def tree():
        return OrderedDict()
    
    root = tree()
    for keys in multiindex:
        current_level = root
        for key in keys:
            if key not in current_level:
                current_level[key] = tree()
            current_level = current_level[key]

    def convert_to_dict(d):
        if isinstance(d, OrderedDict):
            d = {k: convert_to_dict(v) for k, v in d.items()}
        return d
    return convert_to_dict(root)


def postprocess_tb(df, drop=False):
    if drop:
        # Track if index was originally a simple RangeIndex (no semantic meaning)
        # dropna(how='all') can turn RangeIndex into Int64Index by introducing gaps,
        # which would incorrectly trigger the "preserve row index" logic below.
        was_range_index = isinstance(df.index, pd.RangeIndex)
        
        # Drop rows where all data columns are empty (exclude _src_row from the check)
        # _src_row is always non-null, so including it would prevent any row from being dropped.
        src_row_cols = [c for c in df.columns
                        if (isinstance(c, tuple) and c[0] == '_src_row') or c == '_src_row']
        if src_row_cols:
            data_cols = [c for c in df.columns if c not in src_row_cols]
            mask = df[data_cols].isna().all(axis=1)
            df = df[~mask]
        else:
            df = df.dropna(how='all')
        
        # If index was originally RangeIndex, re-number it to avoid gaps
        if was_range_index:
            df = df.reset_index(drop=True)
        
        # Drop columns that are all empty AND have no meaningful header
        # A column with a valid header should be preserved even if data is empty
        cols_to_drop = []
        for col_idx, col in enumerate(df.columns):
            # Check if all data values are NaN
            # Use iloc to avoid ambiguity when MultiIndex has duplicate tuple keys
            if df.iloc[:, col_idx].isna().all():
                # Check if the column header is meaningful
                # For MultiIndex: check if any level has a non-empty meaningful value
                # For simple index: check if the header is not None/empty
                has_meaningful_header = False
                if isinstance(col, tuple):
                    # MultiIndex column - check if any level has meaningful content
                    for level in col:
                        if level and str(level).strip() and str(level).strip() not in ['None', 'nan', 'NaN']:
                            has_meaningful_header = True
                            break
                else:
                    # Simple column name
                    if col and str(col).strip() and str(col).strip() not in ['None', 'nan', 'NaN']:
                        has_meaningful_header = True
                
                # Only drop if header is not meaningful
                if not has_meaningful_header:
                    cols_to_drop.append(col_idx)
        
        if cols_to_drop:
            # Use positional indices to drop columns safely (avoids duplicate MultiIndex key issues)
            cols_to_keep = [i for i in range(len(df.columns)) if i not in cols_to_drop]
            df = df.iloc[:, cols_to_keep]
        
        logger.debug(f"Dropped {len(cols_to_drop)} empty columns")
        
        # Preserve meaningful row index (header columns) as regular columns
        # Only drop=True if it's a simple RangeIndex (no semantic meaning)
        if not isinstance(df.index, pd.RangeIndex):
            # Remember if columns were MultiIndex before reset
            was_multiindex = isinstance(df.columns, pd.MultiIndex)
            n_levels = df.columns.nlevels if was_multiindex else 1
            
            # Avoid name collision: if the index name already exists as a column,
            # clear it before reset_index to prevent "cannot insert X, already exists"
            existing_col_names = {str(c) for c in df.columns}
            if isinstance(df.index, pd.MultiIndex):
                new_names = [None if n is not None and str(n) in existing_col_names else n
                             for n in df.index.names]
                df.index.names = new_names
            elif hasattr(df.index, 'name') and df.index.name is not None:
                if str(df.index.name) in existing_col_names:
                    df.index.name = None
            
            df = df.reset_index()  # Converts index to columns
            
            # Clean up auto-generated column names like 'index', 'level_0', 'level_1'
            # For MultiIndex columns, we need to preserve the structure
            if was_multiindex:
                # Build new column tuples for the index columns
                new_cols = []
                for col in df.columns:
                    if isinstance(col, str) and (col.startswith('level_') or col == 'index'):
                        # Create a tuple with empty strings to match MultiIndex levels
                        new_cols.append(tuple([''] * n_levels))
                    else:
                        new_cols.append(col)
                df.columns = pd.MultiIndex.from_tuples(new_cols)
            else:
                # For simple columns
                new_cols = []
                for col in df.columns:
                    if isinstance(col, str) and (col.startswith('level_') or col == 'index'):
                        new_cols.append('')
                    else:
                        new_cols.append(col)
                df.columns = new_cols
        else:
            df.reset_index(drop=True, inplace=True)

    # Clean column names - preserve MultiIndex structure if present
    if isinstance(df.columns, pd.MultiIndex):
        # For MultiIndex, clean each level's values while preserving structure
        new_levels = []
        for level_idx in range(df.columns.nlevels):
            level_vals = df.columns.get_level_values(level_idx)
            cleaned = [str(v).replace('\n', '') if v is not None else '' for v in level_vals]
            new_levels.append(cleaned)
        df.columns = pd.MultiIndex.from_arrays(new_levels, names=df.columns.names)
        # Also handle 'Unnamed' in MultiIndex
        new_levels = []
        for level_idx in range(df.columns.nlevels):
            level_vals = df.columns.get_level_values(level_idx)
            cleaned = [np.nan if 'Unnamed' in str(v) else v for v in level_vals]
            new_levels.append(cleaned)
        df.columns = pd.MultiIndex.from_arrays(new_levels, names=df.columns.names)
    else:
        df.columns = [str(col).replace('\n', '') for col in df.columns] # Replace '\n' in column headers
        df.columns = [np.nan if 'Unnamed' in str(col) else col for col in df.columns] # Replace Unnamed with nan
    df = df.map(lambda x: x.replace('\n', '') if isinstance(x, str) else x) # Replace '\n' in each cell
    df = process_datetime_cells(df)
    return df


def process_datetime_cells(df):
    df = df.copy()
    def convert(x):
        if isinstance(x, (pd.Timestamp, datetime.datetime)):
            return x.strftime("%Y-%m-%d %H:%M:%S")
        return x
    return df.apply(lambda col: col.map(convert))


def process_duplicate_cols(columns):
    col_count = {}
    new_columns = []
    for col in columns:
        if col in col_count:
            new_columns.append(f"{col}>>>{col_count[col]}")
            col_count[col] += 1
        else:
            new_columns.append(col)
            col_count[col] = 1
    return new_columns


def format_tb_scope(df, num):
    if len(df) > int(num*3+1):
        # Get head and tail rows
        head_df = df.head(num)
        tail_df = df.tail(num)
        # Middle portion excluding head and tail
        middle_df = df.iloc[num:len(df)-num]

        if len(middle_df) >= num:
            mid_sample_df = middle_df.sample(n=num, random_state=42)
        else:  # If middle has less than num rows, take all
            mid_sample_df = middle_df
        scope_df = pd.concat(objs=[head_df, mid_sample_df, tail_df], ignore_index=True)
    else:
        scope_df = df
    scope_df = scope_df.applymap(lambda x: str(x).strip() if pd.notnull(x) else x)
    scope_str = df2html(scope_df)
    return scope_str


def parse_xlsx(file_path, file_name, output_dir, baseurl, base_llm_paras=None, window_h=10, relative_root=None, use_precision_mode=True, include_hidden_sheets=False):
    """
    Parse Excel file and extract table content.
    
    Args:
        file_path: Path or URL to the Excel file
        file_name: Display name for the file
        output_dir: Directory to save extracted tables
        baseurl: Base URL for file loading
        base_llm_paras: LLM parameters for summarization
        window_h: Window size for table scope
        relative_root: Root path for relative paths
        use_precision_mode: If True, use openpyxl merged cell metadata for accurate
                           header detection. If False, use LLM/heuristic mode.
                           Default is True for better accuracy.
        include_hidden_sheets: If True, parse hidden/very-hidden sheets. Default False.
    
    Returns:
        DataFrame with parsed table information
    """
    split_char = settings.SPLIT_CHAR or "/"
    time_stamp = get_str_time()
    df_list = []

    table_data = load_file_bytes(file_path, file_url=baseurl)
    table_stream = io.BytesIO(table_data)
    
    tb_dir = os.path.join(output_dir, "tables")
    os.makedirs(tb_dir, exist_ok=True)
    all_tb_paths = []
    exist_sheets = []

    if use_precision_mode:
        # PRECISION MODE: Use openpyxl metadata for accurate header detection
        logger.info("Using precision mode for Excel header detection")
        try:
            sheets_dict = parse_headers_from_excel(table_stream, include_hidden_sheets=include_hidden_sheets)
            precision_mode_active = True
        except Exception as e:
            logger.warning(f"Precision mode failed, falling back to legacy mode: {e}")
            table_stream.seek(0)  # Reset stream position
            sheets_dict = pd.read_excel(table_stream, sheet_name=None)
            precision_mode_active = False
    else:
        # LEGACY MODE: Use pandas read_excel + LLM/heuristic header detection
        sheets_dict = pd.read_excel(table_stream, sheet_name=None)
        precision_mode_active = False

    all_sheets = sheets_dict.items()

    for sheet_name, sheet_content in all_sheets:
        sheet_name = sheet_name.strip()
        if sheet_name in exist_sheets:
            sheet_name = sheet_name + str(len(exist_sheets))
        else:
            exist_sheets.append(sheet_name)
        
        sheet_tbs = [sheet_content]
        for tb in sheet_tbs:
            try:
                tb = postprocess_tb(tb, drop=True)
                if len(tb) == 0 or tb.empty or tb.isna().all().all():
                    continue

                # In precision mode, headers are already correctly set by parse_headers_from_excel
                # In legacy mode, use LLM/heuristic header parsing
                if not precision_mode_active:
                    tb = parse_headers(tb, paras=base_llm_paras)
                
                # Get row header column count from DataFrame attrs (set in parse_headers_from_excel)
                row_header_cols = tb.attrs.get('row_header_cols', 0)
                
                tb_paths, tb_strs = parse_tb_contents(tb, parent_dic={file_name: {sheet_name: {}}}, file_name=file_name, sheet_name=sheet_name, row_header_cols=row_header_cols)
                tb_keywords = parse_tb_keywords(tb)

                if base_llm_paras['summary_table']:
                    summary_context = format_tb_scope(tb, window_h)
                    summary_context = f"Table columns:\n{tb_keywords}\n\nFirst {window_h} rows:\n{summary_context}"
                    from app.services.document_parser.txt_parser import extract_summary_with_title
                    llm_title, llm_summary = extract_summary_with_title(summary_context, summary_len=100)
                    tb_summary = llm_summary if llm_summary else tb_keywords
                else:
                    llm_title = None
                    tb_summary = tb_keywords

                # Use LLM title for filename when available, fallback to sheet_name
                effective_name = llm_title if llm_title else sheet_name
                tb_name = remove_spaces('table-' + effective_name) + '.html'
                tb_path = os.path.join(tb_dir, tb_name)
                soup = BeautifulSoup(tb_strs, features='html.parser')
                tb_html_str = soup.prettify()
                with open(tb_path, 'w', encoding='utf-8') as f:
                    f.write(tb_html_str)

                tb_id = 'TABLE_' + gen_str_codes(tb_strs) + '_TABLE'
                tb_bottom_content = f"{tb_id}\nTable summary:\n{tb_summary}\nMain columns:\n{tb_keywords}"
                
                know_id = gen_str_codes(tb_bottom_content + str(uuid.uuid4()))
                bottom_tokens = tokenize2stw_remove([tb_bottom_content], base_llm_paras['stopwords'])

                all_tb_paths.extend(tb_paths)
                # Use relative path for tables: "tables/xxx.html"
                relative_tb_path = f"tables/{tb_name}"
                df_list.append([tb_bottom_content, relative_tb_path, tb_id, len(tb_strs), tb_keywords, tb_summary, know_id, bottom_tokens, "", time_stamp, ""])

            except KnowhereException:
                raise
            except Exception as e:
                logger.error(f"Table parsing failed: {e}")
                raise TableParsingException(
                    user_message="Failed to parse Excel table content",
                    reason="TABLE_PROCESSING_FAILED",
                    internal_message=str(e),
                    original_exception=e
                )

    table_df = pd.DataFrame(df_list, columns=settings.ALL_DF_COLS.split(','))
    table_df = process_dup_paths_df(table_df)
    return table_df
