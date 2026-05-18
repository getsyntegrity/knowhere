from __future__ import annotations

import datetime
import io
from typing import List, Optional, Tuple, Union
from typing import cast

import openpyxl
import pandas as pd
from pandas._typing import Axes
from loguru import logger

from shared.core.exceptions.domain_exceptions import TableParsingException


def parse_excel_structure(
    file_source: Union[str, io.BytesIO],
    sheet_name: Optional[str] = None,
    split_subtables: bool = True,
    include_hidden_sheets: bool = False,
) -> dict[str, pd.DataFrame]:
    try:
        if isinstance(file_source, str):
            workbook = openpyxl.load_workbook(file_source, data_only=True)
        else:
            file_source.seek(0)
            workbook = openpyxl.load_workbook(file_source, data_only=True)

        results: dict[str, pd.DataFrame] = {}
        sheets_to_parse = [sheet_name] if sheet_name else workbook.sheetnames

        for selected_sheet_name in sheets_to_parse:
            if selected_sheet_name not in workbook.sheetnames:
                logger.warning(
                    f"Sheet '{selected_sheet_name}' not found in workbook, skipping"
                )
                continue

            worksheet = workbook[selected_sheet_name]

            if not include_hidden_sheets and worksheet.sheet_state != "visible":
                logger.info(
                    f"Sheet '{selected_sheet_name}' is hidden "
                    f"(state={worksheet.sheet_state}), skipping"
                )
                continue

            if worksheet.max_row is None or worksheet.max_row == 0:
                logger.debug(f"Sheet '{selected_sheet_name}' is empty, skipping")
                continue

            merged_ranges = list(worksheet.merged_cells.ranges)
            logger.debug(
                f"Sheet '{selected_sheet_name}': found {len(merged_ranges)} merged cell ranges"
            )

            if split_subtables:
                subtable_regions = _split_sheet_recursive(
                    worksheet,
                    (1, worksheet.max_row),
                    (1, worksheet.max_column or 1),
                    merged_ranges,
                )
                before_count = len(subtable_regions)
                subtable_regions = _merge_small_subtables(worksheet, subtable_regions)
                if len(subtable_regions) != before_count:
                    logger.info(
                        f"Sheet '{selected_sheet_name}': merged {before_count} subtables → "
                        f"{len(subtable_regions)} "
                        f"(absorbed {before_count - len(subtable_regions)} small fragments)"
                    )
                logger.debug(
                    f"Sheet '{selected_sheet_name}': {len(subtable_regions)} subtables after merge"
                )

                for index, (row_range, col_range) in enumerate(subtable_regions):
                    result = _parse_subtable(
                        worksheet,
                        row_range,
                        col_range,
                        merged_ranges,
                    )
                    dataframe = result["df"]
                    dataframe.attrs["row_header_cols"] = len(result["header_cols"])
                    key = selected_sheet_name if index == 0 else f"{selected_sheet_name}_{index + 1}"
                    logger.debug(
                        f"Subtable '{key}': rows={row_range}, cols={col_range}, "
                        f"header_rows={result['header_rows']}, header_cols={result['header_cols']}"
                    )
                    results[key] = dataframe
            else:
                row_range = (1, worksheet.max_row)
                col_range = (1, worksheet.max_column or 1)
                result = _parse_subtable(worksheet, row_range, col_range, merged_ranges)
                dataframe = result["df"]
                dataframe.attrs["row_header_cols"] = len(result["header_cols"])
                logger.debug(
                    f"Sheet '{selected_sheet_name}': header_rows={result['header_rows']}, "
                    f"header_cols={result['header_cols']}, "
                    f"fallback_col={result['fallback_col_header']}, "
                    f"fallback_row={result['fallback_row_header']}"
                )
                results[selected_sheet_name] = dataframe

        workbook.close()
        return results
    except Exception as exc:
        logger.error(f"Error parsing Excel with precision mode: {exc}")
        raise TableParsingException(
            user_message="Failed to parse Excel file headers",
            reason="EXCEL_PRECISION_PARSE_FAILED",
            internal_message=str(exc),
            original_exception=exc,
        ) from exc


DATA_TYPES_TO_EXCLUDE = (int, float, datetime.datetime)


def _get_merged_cell_value(ws, row: int, col: int, merged_ranges: list):
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return ws.cell(row, col).value


def _get_unique_cells_in_row(
    ws,
    row: int,
    col_range: Tuple[int, int],
    merged_ranges: list,
) -> List[dict]:
    col_start, col_end = col_range
    cells = []
    visited_cols = set()

    for col in range(col_start, col_end + 1):
        if col in visited_cols:
            continue

        in_merge = False
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                value = ws.cell(merged_range.min_row, merged_range.min_col).value
                merge_col_end = min(merged_range.max_col, col_end)

                for merged_col in range(merged_range.min_col, merge_col_end + 1):
                    visited_cols.add(merged_col)

                cells.append(
                    {
                        "col_start": merged_range.min_col,
                        "col_end": merge_col_end,
                        "value": value,
                        "is_merged": True,
                    }
                )
                in_merge = True
                break

        if not in_merge:
            value = ws.cell(row, col).value
            cells.append(
                {"col_start": col, "col_end": col, "value": value, "is_merged": False}
            )
            visited_cols.add(col)

    return cells


def _get_unique_cells_in_col(
    ws,
    col: int,
    row_range: Tuple[int, int],
    merged_ranges: list,
) -> List[dict]:
    row_start, row_end = row_range
    cells = []
    visited_rows = set()

    for row in range(row_start, row_end + 1):
        if row in visited_rows:
            continue

        in_merge = False
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                value = ws.cell(merged_range.min_row, merged_range.min_col).value
                merge_row_end = min(merged_range.max_row, row_end)

                for merged_row in range(merged_range.min_row, merge_row_end + 1):
                    visited_rows.add(merged_row)

                cells.append(
                    {
                        "row_start": merged_range.min_row,
                        "row_end": merge_row_end,
                        "value": value,
                        "is_merged": True,
                    }
                )
                in_merge = True
                break

        if not in_merge:
            value = ws.cell(row, col).value
            cells.append(
                {"row_start": row, "row_end": row, "value": value, "is_merged": False}
            )
            visited_rows.add(row)

    return cells


def _is_candidate_header_row(
    ws,
    row: int,
    col_range: Tuple[int, int],
    merged_ranges: list,
    exclude_types: tuple = DATA_TYPES_TO_EXCLUDE,
) -> bool:
    cells = _get_unique_cells_in_row(ws, row, col_range, merged_ranges)

    has_any_value = False
    for cell in cells:
        value = cell["value"]
        if value is None:
            continue
        has_any_value = True

        if isinstance(value, bool):
            continue
        if isinstance(value, exclude_types):
            return False

    return has_any_value


def _is_candidate_header_col(
    ws,
    col: int,
    row_range: Tuple[int, int],
    merged_ranges: list,
    exclude_types: tuple = DATA_TYPES_TO_EXCLUDE,
) -> bool:
    cells = _get_unique_cells_in_col(ws, col, row_range, merged_ranges)

    has_any_value = False
    for cell in cells:
        value = cell["value"]
        if value is None:
            continue
        has_any_value = True

        if isinstance(value, bool):
            continue
        if isinstance(value, exclude_types):
            return False

    return has_any_value


def _detect_header_regions(
    ws,
    row_range: Tuple[int, int],
    col_range: Tuple[int, int],
    merged_ranges: list,
) -> Tuple[List[int], List[int]]:
    row_start, row_end = row_range
    col_start, col_end = col_range

    header_rows = []
    for row in range(row_start, row_end + 1):
        if _is_candidate_header_row(ws, row, col_range, merged_ranges):
            header_rows.append(row)
        else:
            break

    data_row_start = header_rows[-1] + 1 if header_rows else row_start
    if data_row_start > row_end:
        return header_rows, []

    header_cols = []
    data_row_range = (data_row_start, row_end)
    for col in range(col_start, col_end + 1):
        if _is_candidate_header_col(ws, col, data_row_range, merged_ranges):
            header_cols.append(col)
        else:
            break

    return header_rows, header_cols


def _build_column_multiindex(
    ws,
    header_rows: List[int],
    col_range: Tuple[int, int],
    merged_ranges: list,
) -> Union[pd.Index, pd.MultiIndex]:
    col_start, col_end = col_range
    levels = []

    for row in header_rows:
        row_values = []
        for col in range(col_start, col_end + 1):
            value = _get_merged_cell_value(ws, row, col, merged_ranges)
            row_values.append(str(value).strip() if value else "")
        levels.append(row_values)

    for index, level in enumerate(levels):
        filled = []
        last = ""
        for value in level:
            if value:
                last = value
            filled.append(last if last else value)
        levels[index] = filled

    if len(levels) == 1:
        return pd.Index(levels[0])
    return pd.MultiIndex.from_arrays(levels)


def _build_row_multiindex(
    ws,
    header_cols: List[int],
    row_range: Tuple[int, int],
    merged_ranges: list,
    header_rows: List[int] | None = None,
) -> Union[pd.Index, pd.MultiIndex]:
    row_start, row_end = row_range
    levels = []
    names = []

    for col in header_cols:
        col_values = []
        for row in range(row_start, row_end + 1):
            value = _get_merged_cell_value(ws, row, col, merged_ranges)
            col_values.append(str(value).strip() if value else "")
        levels.append(col_values)

        if header_rows:
            name_row = header_rows[-1]
            name_value = _get_merged_cell_value(ws, name_row, col, merged_ranges)
            names.append(str(name_value).strip() if name_value else None)
        else:
            names.append(None)

    for index, level in enumerate(levels):
        filled = []
        last = ""
        for value in level:
            if value:
                last = value
            filled.append(last if last else value)
        levels[index] = filled

    if len(levels) == 1:
        row_index = pd.Index(levels[0])
        row_index.name = names[0] if names else None
        return row_index
    return pd.MultiIndex.from_arrays(levels, names=names)


def _parse_subtable(
    ws,
    row_range: Tuple[int, int],
    col_range: Tuple[int, int],
    merged_ranges: list,
) -> dict:
    row_start, row_end = row_range
    col_start, col_end = col_range

    header_rows, header_cols = _detect_header_regions(
        ws, row_range, col_range, merged_ranges
    )

    total_rows = row_end - row_start + 1
    total_cols = col_end - col_start + 1
    fallback_col_header = len(header_rows) == total_rows
    fallback_row_header = len(header_cols) == total_cols

    if fallback_col_header:
        data_row_start = row_start
        columns = None
    else:
        data_row_start = header_rows[-1] + 1 if header_rows else row_start
        columns = (
            _build_column_multiindex(ws, header_rows, col_range, merged_ranges)
            if header_rows
            else None
        )

    if fallback_row_header:
        data_col_start = col_start
        row_index = None
    else:
        data_col_start = header_cols[-1] + 1 if header_cols else col_start
        row_index = (
            _build_row_multiindex(
                ws, header_cols, (data_row_start, row_end), merged_ranges, header_rows
            )
            if header_cols
            else None
        )

    data = []
    for row in range(data_row_start, row_end + 1):
        row_data = []
        for col in range(data_col_start, col_end + 1):
            value = _get_merged_cell_value(ws, row, col, merged_ranges)
            row_data.append(value)
        data.append(row_data)

    if columns is not None and header_cols and not fallback_row_header:
        columns = cast(Axes, columns[len(header_cols) :])

    dataframe = pd.DataFrame(data, columns=columns, index=row_index)

    excel_row_numbers = list(range(data_row_start, row_end + 1))
    if isinstance(dataframe.columns, pd.MultiIndex):
        level_count = dataframe.columns.nlevels
        src_row_key = tuple(["_src_row"] + [""] * (level_count - 1))
        dataframe[src_row_key] = excel_row_numbers
    else:
        dataframe["_src_row"] = excel_row_numbers

    return {
        "df": dataframe,
        "header_rows": header_rows if not fallback_col_header else [],
        "header_cols": header_cols if not fallback_row_header else [],
        "fallback_col_header": fallback_col_header,
        "fallback_row_header": fallback_row_header,
    }


def _find_effective_range(
    ws,
    row_range: Tuple[int, int],
    col_range: Tuple[int, int],
) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    row_start, row_end = row_range
    col_start, col_end = col_range

    effective_row_start = None
    effective_row_end = None
    effective_col_start = None
    effective_col_end = None

    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            if ws.cell(row, col).value is not None:
                if effective_row_start is None:
                    effective_row_start = row
                effective_row_end = row
                if effective_col_start is None or col < effective_col_start:
                    effective_col_start = col
                if effective_col_end is None or col > effective_col_end:
                    effective_col_end = col

    if effective_row_start is None:
        return ((row_start, row_start), (col_start, col_start))

    if (
        effective_row_end is None
        or effective_col_start is None
        or effective_col_end is None
    ):
        return ((row_start, row_start), (col_start, col_start))

    return (
        (effective_row_start, effective_row_end),
        (effective_col_start, effective_col_end),
    )


def _is_true_separator_row(
    ws,
    row: int,
    effective_col_range: Tuple[int, int],
    merged_ranges: list | None = None,
) -> bool:
    col_start, col_end = effective_col_range
    merged_ranges = merged_ranges or []

    for col in range(col_start, col_end + 1):
        if ws.cell(row, col).value is not None:
            return False
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return False
    return True


def _is_true_separator_col(
    ws,
    col: int,
    effective_row_range: Tuple[int, int],
    merged_ranges: list | None = None,
) -> bool:
    row_start, row_end = effective_row_range
    merged_ranges = merged_ranges or []

    for row in range(row_start, row_end + 1):
        if ws.cell(row, col).value is not None:
            return False
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return False
    return True


def _find_separator_groups(items: List[int]) -> List[List[int]]:
    if not items:
        return []

    groups = []
    current_group = [items[0]]

    for index in range(1, len(items)):
        if items[index] == items[index - 1] + 1:
            current_group.append(items[index])
        else:
            groups.append(current_group)
            current_group = [items[index]]

    groups.append(current_group)
    return groups


def _split_sheet_recursive(
    ws,
    row_range: Tuple[int, int],
    col_range: Tuple[int, int],
    merged_ranges: list | None = None,
    min_rows: int = 2,
    min_cols: int = 2,
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    row_start, row_end = row_range
    col_start, col_end = col_range
    merged_ranges = merged_ranges or []

    (effective_row_start, effective_row_end), (
        effective_col_start,
        effective_col_end,
    ) = _find_effective_range(ws, row_range, col_range)

    if (
        effective_row_end - effective_row_start + 1 < min_rows
        or effective_col_end - effective_col_start + 1 < min_cols
    ):
        if effective_row_start is not None:
            return [
                (
                    (effective_row_start, effective_row_end),
                    (effective_col_start, effective_col_end),
                )
            ]
        return []

    separator_rows = []
    for row in range(effective_row_start + 1, effective_row_end):
        if _is_true_separator_row(
            ws, row, (effective_col_start, effective_col_end), merged_ranges
        ):
            separator_rows.append(row)

    separator_cols = []
    for col in range(effective_col_start + 1, effective_col_end):
        if _is_true_separator_col(
            ws, col, (effective_row_start, effective_row_end), merged_ranges
        ):
            separator_cols.append(col)

    row_groups = _find_separator_groups(separator_rows)
    col_groups = _find_separator_groups(separator_cols)

    should_split_rows = len(row_groups) > 0 and (
        len(col_groups) == 0 or len(row_groups) <= len(col_groups)
    )
    should_split_cols = len(col_groups) > 0 and not should_split_rows

    if should_split_rows:
        subtables = []
        previous_end = effective_row_start
        for group in row_groups:
            if group[0] > previous_end:
                sub_result = _split_sheet_recursive(
                    ws,
                    (previous_end, group[0] - 1),
                    (effective_col_start, effective_col_end),
                    merged_ranges,
                    min_rows,
                    min_cols,
                )
                subtables.extend(sub_result)
            previous_end = group[-1] + 1
        if previous_end <= effective_row_end:
            sub_result = _split_sheet_recursive(
                ws,
                (previous_end, effective_row_end),
                (effective_col_start, effective_col_end),
                merged_ranges,
                min_rows,
                min_cols,
            )
            subtables.extend(sub_result)
        return subtables

    if should_split_cols:
        subtables = []
        previous_end = effective_col_start
        for group in col_groups:
            if group[0] > previous_end:
                sub_result = _split_sheet_recursive(
                    ws,
                    (effective_row_start, effective_row_end),
                    (previous_end, group[0] - 1),
                    merged_ranges,
                    min_rows,
                    min_cols,
                )
                subtables.extend(sub_result)
            previous_end = group[-1] + 1
        if previous_end <= effective_col_end:
            sub_result = _split_sheet_recursive(
                ws,
                (effective_row_start, effective_row_end),
                (previous_end, effective_col_end),
                merged_ranges,
                min_rows,
                min_cols,
            )
            subtables.extend(sub_result)
        return subtables

    return [((effective_row_start, effective_row_end), (effective_col_start, effective_col_end))]


def _count_non_empty_cells(
    ws,
    row_range: Tuple[int, int],
    col_range: Tuple[int, int],
) -> int:
    count = 0
    for row in range(row_range[0], row_range[1] + 1):
        for col in range(col_range[0], col_range[1] + 1):
            if ws.cell(row, col).value is not None:
                count += 1
    return count


def _merge_small_subtables(
    ws,
    subtables: List[Tuple[Tuple[int, int], Tuple[int, int]]],
    min_cells: int = 4,
) -> List[Tuple[Tuple[int, int], Tuple[int, int]]]:
    if len(subtables) <= 1:
        return subtables

    items = []
    for row_range, col_range in subtables:
        count = _count_non_empty_cells(ws, row_range, col_range)
        items.append({"rr": row_range, "cr": col_range, "cells": count})

    changed = True
    while changed and len(items) > 1:
        changed = False

        min_index = None
        for index, item in enumerate(items):
            if item["cells"] < min_cells:
                if min_index is None or item["cells"] < items[min_index]["cells"]:
                    min_index = index

        if min_index is None:
            break

        source = items[min_index]
        best_index = None
        best_distance = float("inf")
        for index, target in enumerate(items):
            if index == min_index:
                continue
            row_gap = max(
                0,
                target["rr"][0] - source["rr"][1] - 1,
                source["rr"][0] - target["rr"][1] - 1,
            )
            col_gap = max(
                0,
                target["cr"][0] - source["cr"][1] - 1,
                source["cr"][0] - target["cr"][1] - 1,
            )
            distance = row_gap + col_gap
            if distance < best_distance or (
                best_index is not None
                and distance == best_distance
                and target["cells"] > items[best_index]["cells"]
            ):
                best_distance = distance
                best_index = index

        if best_index is None:
            break

        target = items[best_index]
        merged_row_range = (
            min(source["rr"][0], target["rr"][0]),
            max(source["rr"][1], target["rr"][1]),
        )
        merged_col_range = (
            min(source["cr"][0], target["cr"][0]),
            max(source["cr"][1], target["cr"][1]),
        )
        items[best_index] = {
            "rr": merged_row_range,
            "cr": merged_col_range,
            "cells": source["cells"] + target["cells"],
        }

        logger.debug(
            f"Merged small fragment (rows={source['rr']}, cols={source['cr']}, "
            f"cells={source['cells']}) into neighbor (rows={target['rr']}, cols={target['cr']})"
        )

        del items[min_index]
        changed = True

    return [(item["rr"], item["cr"]) for item in items]
